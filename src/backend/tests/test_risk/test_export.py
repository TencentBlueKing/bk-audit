# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making
蓝鲸智云 - 审计中心 (BlueKing - Audit Center) available.
Copyright (C) 2023 THL A29 Limited,
a Tencent company. All rights reserved.
Licensed under the MIT License (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at https://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing,
software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND,
either express or implied. See the License for the
specific language governing permissions and limitations under the License.
We undertake not to change the open source license (MIT license) applicable
to the current version of the project delivered to anyone in the future.
"""
import base64
import datetime
import io
from unittest import mock
from urllib.parse import unquote

import openpyxl
from bk_resource import resource
from django.conf import settings
from django.test import override_settings
from rest_framework.response import Response

from core.utils.time import mstimestamp_to_date_string
from services.web.risk.constants import (
    RAW_EVENT_ID_REMARK,
    RiskExportField,
    RiskStatus,
    RiskViewType,
)
from services.web.risk.handlers.risk_export_service import RiskExportService
from services.web.risk.models import Risk
from services.web.risk.resources import ListEvent
from services.web.risk.serializers import (
    RetrieveRiskStrategyInfoResponseSerializer,
    RiskExportReqSerializer,
)
from services.web.risk.tasks import export_risks_to_mail
from services.web.scene.constants import ResourceVisibilityType
from services.web.scene.filters import BindingMetadataHelper
from services.web.scene.models import Scene
from services.web.strategy_v2.constants import RiskLevel
from services.web.strategy_v2.models import Strategy
from tests.base import TestCase


class TestRiskExport(TestCase):
    """
    测试风险导出
    """

    def setUp(self):
        super().setUp()
        self.maxDiff = None
        self.scene = Scene.objects.create(name="risk-export-scene")
        self.other_scene = Scene.objects.create(name="risk-export-other-scene")
        # 创建策略
        self.strategy_1 = Strategy.objects.create(
            strategy_id=1,
            strategy_name="Strategy A",
            risk_level=RiskLevel.HIGH,
            event_data_field_configs=[
                {"field_name": "user", "display_name": "Username", "duplicate_field": False},
                {"field_name": "action", "display_name": "Action", "duplicate_field": False},
            ],
        )
        self.strategy_2 = Strategy.objects.create(
            strategy_id=2,
            strategy_name="Strategy B",
            risk_level=RiskLevel.MIDDLE,
            event_data_field_configs=[{"field_name": "ip", "display_name": "Source IP", "duplicate_field": False}],
        )
        BindingMetadataHelper.create_resource_binding(
            resource_id=str(self.strategy_1.strategy_id),
            resource_type=ResourceVisibilityType.STRATEGY,
            scene_id=self.scene.scene_id,
        )
        BindingMetadataHelper.create_resource_binding(
            resource_id=str(self.strategy_2.strategy_id),
            resource_type=ResourceVisibilityType.STRATEGY,
            scene_id=self.scene.scene_id,
        )
        # 创建风险
        self.risk_1 = Risk.objects.create(
            risk_id="risk001",
            title="Risk 1 Title",
            strategy=self.strategy_1,
            status=RiskStatus.NEW,
            event_time=datetime.datetime(2023, 1, 1, 10, 0, 0),
            event_end_time=datetime.datetime(2023, 1, 1, 11, 0, 0),
            event_type=["login"],
            operator=["admin"],
            current_operator=["admin"],
            notice_users=["user_a"],
        )
        self.risk_2 = Risk.objects.create(
            risk_id="risk002",
            title="Risk 2 Title",
            strategy=self.strategy_1,
            status=RiskStatus.AWAIT_PROCESS,
            event_time=datetime.datetime(2023, 1, 2, 10, 0, 0),
            event_end_time=datetime.datetime(2023, 1, 2, 11, 0, 0),
            event_type=["delete"],
            operator=["user_b"],
            current_operator=["user_b"],
            notice_users=["user_c"],
        )
        self.risk_3 = Risk.objects.create(
            risk_id="risk003",
            title="Risk 3 Title",
            strategy=self.strategy_2,
            status=RiskStatus.CLOSED,
            event_time=datetime.datetime(2023, 1, 3, 10, 0, 0),
            event_end_time=datetime.datetime(2023, 1, 3, 11, 0, 0),
            event_type=["download"],
            operator=["user_d"],
            current_operator=[],
            notice_users=[],
        )
        # 风险4没有关联事件
        self.risk_4 = Risk.objects.create(
            risk_id="risk004",
            title="Risk 4 Title",
            strategy=self.strategy_2,
            status=RiskStatus.NEW,
            event_time=datetime.datetime(2023, 1, 4, 10, 0, 0),
            event_end_time=datetime.datetime(2023, 1, 4, 11, 0, 0),
            event_type=["upload"],
            operator=["user_e"],
            current_operator=["user_e"],
            notice_users=[],
        )

    def mock_get_event_list(self, risk_id, **kwargs):
        return [
            {
                "results": [
                    {"event_data": {"user": "alice", "action": "login_success"}},
                    {"event_data": {"user": "alice", "action": "login_fail"}},
                ]
            },
            {"results": [{"event_data": {"user": "bob", "action": "delete_file"}}]},
            {"results": [{"event_data": {"ip": "127.0.0.1"}}, {"event_data": {"ip": "192.168.1.1"}}]},
            {"results": []},
        ]

    def test_risk_export_serializer_allows_async_limit(self):
        max_count = settings.RISK_EXPORT_ASYNC_MAX_COUNT
        serializer = RiskExportReqSerializer(data={"risk_ids": [f"risk-{index}" for index in range(max_count)]})

        self.assertTrue(serializer.is_valid(), serializer.errors)
        self.assertEqual(serializer.fields["risk_ids"].max_length, max_count)
        self.assertEqual(len(serializer.validated_data["risk_ids"]), max_count)

    def test_risk_export_serializer_blocks_over_async_limit(self):
        serializer = RiskExportReqSerializer(
            data={"risk_ids": [f"risk-{index}" for index in range(settings.RISK_EXPORT_ASYNC_MAX_COUNT + 1)]}
        )

        self.assertFalse(serializer.is_valid())
        self.assertIn("risk_ids", serializer.errors)

    @override_settings(RISK_EXPORT_SYNC_MAX_COUNT=1, RISK_EXPORT_ASYNC_MAX_COUNT=10)
    @mock.patch("services.web.risk.resources.risk.export_risks_to_mail")
    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    def test_risk_export_auto_uses_async_when_over_sync_limit(self, mock_load_authed_risks, mock_task):
        mock_load_authed_risks.return_value = Risk.objects.filter(
            risk_id__in=[self.risk_1.risk_id, self.risk_2.risk_id]
        )
        mock_task.apply_async.return_value = mock.Mock(id="task-risk-export-1")

        resp = resource.risk.risk_export.request(
            risk_ids=[self.risk_1.risk_id, self.risk_2.risk_id],
            risk_view_type=RiskViewType.ALL.value,
        )

        self.assertIsInstance(resp, Response)
        self.assertEqual(resp.status_code, 202)
        self.assertEqual(resp.data["export_type"], "async")
        self.assertEqual(resp.data["task_id"], "task-risk-export-1")
        self.assertEqual(resp.data["total"], 2)
        self.assertEqual(resp.data["notice_users"], ["admin"])
        mock_task.apply_async.assert_called_once()

    @override_settings(RISK_EXPORT_SYNC_MAX_COUNT=300, RISK_EXPORT_ASYNC_MAX_COUNT=400)
    @mock.patch("services.web.risk.resources.risk.export_risks_to_mail")
    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export_large_request_only_schedules_async(
        self, mock_get_event_list, mock_load_authed_risks, mock_task
    ):
        risk_ids = [f"bulk-risk-{index:03d}" for index in range(301)]
        Risk.objects.bulk_create(
            [
                Risk(
                    risk_id=risk_id,
                    title=f"Bulk Risk {index}",
                    strategy=self.strategy_1,
                    status=RiskStatus.NEW,
                    event_time=datetime.datetime(2023, 1, 6, 10, 0, 0),
                    event_end_time=datetime.datetime(2023, 1, 6, 11, 0, 0),
                )
                for index, risk_id in enumerate(risk_ids)
            ]
        )
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id__in=risk_ids)
        mock_task.apply_async.return_value = mock.Mock(id="task-risk-export-large")

        resp = resource.risk.risk_export.request(risk_ids=risk_ids, risk_view_type=RiskViewType.ALL.value)

        self.assertEqual(resp.status_code, 202)
        self.assertEqual(resp.data["task_id"], "task-risk-export-large")
        self.assertEqual(resp.data["total"], 301)
        mock_get_event_list.assert_not_called()
        task_kwargs = mock_task.apply_async.call_args.kwargs["kwargs"]
        self.assertEqual(task_kwargs["risk_ids"], risk_ids)
        self.assertEqual(task_kwargs["risk_view_type"], RiskViewType.ALL.value)

    @override_settings(RISK_EXPORT_SYNC_MAX_COUNT=2, RISK_EXPORT_ASYNC_MAX_COUNT=10)
    @mock.patch("services.web.risk.resources.risk.export_risks_to_mail")
    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export_uses_sync_when_equal_sync_limit(self, mock_get_event_list, mock_load_authed_risks, mock_task):
        mock_load_authed_risks.return_value = Risk.objects.filter(
            risk_id__in=[self.risk_1.risk_id, self.risk_2.risk_id]
        )
        mock_get_event_list.return_value = [{"results": []}, {"results": []}]

        resp = resource.risk.risk_export.request(
            risk_ids=[self.risk_1.risk_id, self.risk_2.risk_id],
            risk_view_type=RiskViewType.ALL.value,
        )

        self.assertIn("attachment; filename*", resp["Content-Disposition"])
        mock_task.apply_async.assert_not_called()

    @override_settings(RISK_EXPORT_SYNC_MAX_COUNT=300, RISK_EXPORT_ASYNC_MAX_COUNT=10)
    @mock.patch("services.web.risk.resources.risk.export_risks_to_mail")
    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    def test_risk_export_async_always_uses_celery(self, mock_load_authed_risks, mock_task):
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)
        mock_task.apply_async.return_value = mock.Mock(id="task-risk-export-small")

        resp = resource.risk.risk_export_async.request(
            risk_ids=[self.risk_1.risk_id],
            risk_view_type=RiskViewType.ALL.value,
        )

        self.assertIsInstance(resp, Response)
        self.assertEqual(resp.status_code, 202)
        self.assertEqual(resp.data["export_type"], "async")
        self.assertEqual(resp.data["task_id"], "task-risk-export-small")
        self.assertEqual(resp.data["total"], 1)
        mock_task.apply_async.assert_called_once()

    @mock.patch("services.web.risk.resources.risk.export_risks_to_mail")
    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    def test_risk_export_async_blocks_unauthed_risks_before_celery(self, mock_load_authed_risks, mock_task):
        from services.web.risk.exceptions import ExportRiskNoPermission

        mock_load_authed_risks.return_value = Risk.objects.none()

        with self.assertRaises(ExportRiskNoPermission):
            resource.risk.risk_export_async.request(
                risk_ids=[self.risk_1.risk_id],
                risk_view_type=RiskViewType.ALL.value,
            )

        mock_task.apply_async.assert_not_called()

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export_service_builds_file(self, mock_get_event_list, mock_load_authed_risks):
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id__in=[self.risk_1.risk_id])
        mock_get_event_list.return_value = [{"results": []}]

        with self.assertLogs("services.web.risk.handlers.risk_export_service", level="INFO") as log_context:
            export_file = RiskExportService(
                username="admin",
                risk_ids=[self.risk_1.risk_id],
                risk_view_type=RiskViewType.ALL.value,
            ).build_export_file()

        self.assertEqual(export_file.total, 1)
        self.assertIn("审计风险", export_file.filename)
        workbook = openpyxl.load_workbook(io.BytesIO(export_file.file.read()))
        self.assertEqual(workbook.sheetnames, [self.strategy_1.build_sheet_name()])
        logs = "\n".join(log_context.output)
        self.assertIn("[RiskExportService] build export file start", logs)
        self.assertIn("[RiskExportService] build export file finished", logs)

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    @mock.patch("services.web.risk.handlers.risk_export_service.MailSender")
    def test_risk_export_service_send_mail_with_attachment(
        self, mock_mail_sender, mock_get_event_list, mock_load_authed_risks
    ):
        sender = mock_mail_sender.return_value
        sender.send.return_value = {"result": True, "code": 0, "message": "OK"}
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)
        mock_get_event_list.return_value = [{"results": []}]
        export_file = RiskExportService(
            username="admin",
            risk_ids=[self.risk_1.risk_id],
            risk_view_type=RiskViewType.ALL.value,
        ).build_export_file()
        export_file.file.seek(0)
        expected_content = base64.b64encode(export_file.file.read()).decode("utf-8")

        result = RiskExportService(
            username="admin",
            risk_ids=[self.risk_1.risk_id],
            risk_view_type=RiskViewType.ALL.value,
        ).send_mail(export_file=export_file, requested_at="2026-06-22 19:33:36")

        self.assertEqual(result["result"], True)
        mock_mail_sender.assert_called_once()
        sender.send.assert_called_once()
        kwargs = mock_mail_sender.call_args.kwargs
        self.assertEqual(kwargs["receivers"], ["admin"])
        self.assertEqual(kwargs["title"], "【蓝鲸审计中心】风险数据导出结果通知")
        self.assertEqual(kwargs["attachments"][0]["filename"], export_file.filename)
        self.assertEqual(kwargs["attachments"][0]["content"], expected_content)
        self.assertEqual(kwargs["attachments"][0]["type"], "xlsx")
        self.assertEqual(kwargs["attachments"][0]["disposition"], "attachment")
        content_text = str(kwargs["content"])
        self.assertNotIn("【蓝鲸审计中心】风险数据导出结果通知", content_text)
        self.assertIn("您好 admin：", content_text)
        self.assertIn("2026-06-22 19:33:36", content_text)
        self.assertIn(export_file.view_label, content_text)
        self.assertIn("本次共导出风险单 1 条", content_text)
        self.assertIn("详细数据请见附件", content_text)
        self.assertIn("此邮件为系统自动发送，请勿直接回复", content_text)

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export(self, mock_get_event_list, mock_load_authed_risks):
        # Mock
        mock_load_authed_risks.return_value = Risk.objects.all()
        mock_get_event_list.side_effect = self.mock_get_event_list

        # Call resource
        risk_ids = ["risk001", "risk002", "risk003", "risk004"]
        resp = resource.risk.risk_export(risk_ids=risk_ids)

        # Verify response
        self.assertIn("attachment; filename*", resp["Content-Disposition"])
        self.assertEqual(resp["Content-Type"], "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

        # Verify excel content
        workbook = openpyxl.load_workbook(io.BytesIO(b"".join(resp.streaming_content)))

        # 1. Check sheets
        sheet1_name = self.strategy_1.build_sheet_name()
        sheet2_name = self.strategy_2.build_sheet_name()
        self.assertEqual(workbook.sheetnames, [sheet1_name, sheet2_name])

        # 2. Check Sheet 1 (Strategy A)
        sheet1 = workbook[sheet1_name]
        # Check headers
        expected_headers1 = [f.display_name for f in RiskExportField.export_fields()] + ["Username", "Action"]
        actual_headers1 = [cell.value for cell in sheet1[1]]
        self.assertEqual(actual_headers1, expected_headers1)
        # Check data rows
        self.assertEqual(sheet1.max_row, 4)  # 1 header + 3 data rows
        # Create header map
        header_map1 = {cell.value: i for i, cell in enumerate(sheet1[1], 1)}
        # Row 1 (risk001, event 1)
        self.assertEqual(sheet1.cell(row=2, column=header_map1[str(RiskExportField.RISK_ID.label)]).value, "risk001")
        self.assertEqual(sheet1.cell(row=2, column=header_map1["Username"]).value, "alice")
        self.assertEqual(sheet1.cell(row=2, column=header_map1["Action"]).value, "login_success")
        # Row 2 (risk001, event 2)
        self.assertEqual(sheet1.cell(row=3, column=header_map1[str(RiskExportField.RISK_ID.label)]).value, "risk001")
        self.assertEqual(sheet1.cell(row=3, column=header_map1["Username"]).value, "alice")
        self.assertEqual(sheet1.cell(row=3, column=header_map1["Action"]).value, "login_fail")
        # Row 3 (risk002, event 1)
        self.assertEqual(sheet1.cell(row=4, column=header_map1[str(RiskExportField.RISK_ID.label)]).value, "risk002")
        self.assertEqual(sheet1.cell(row=4, column=header_map1["Username"]).value, "bob")
        self.assertEqual(sheet1.cell(row=4, column=header_map1["Action"]).value, "delete_file")

        # 3. Check Sheet 2 (Strategy B)
        sheet2 = workbook[sheet2_name]
        # Check headers
        expected_headers2 = [f.display_name for f in RiskExportField.export_fields()] + ["Source IP"]
        actual_headers2 = [cell.value for cell in sheet2[1]]
        self.assertEqual(actual_headers2, expected_headers2)
        # Check data rows
        self.assertEqual(sheet2.max_row, 4)  # 1 header + 3 data rows
        # Create header map
        header_map2 = {cell.value: i for i, cell in enumerate(sheet2[1], 1)}
        # Row 1 (risk003, event 1)
        self.assertEqual(sheet2.cell(row=2, column=header_map2[str(RiskExportField.RISK_ID.label)]).value, "risk003")
        self.assertEqual(sheet2.cell(row=2, column=header_map2["Source IP"]).value, "127.0.0.1")
        # Row 2 (risk003, event 2)
        self.assertEqual(sheet2.cell(row=3, column=header_map2[str(RiskExportField.RISK_ID.label)]).value, "risk003")
        self.assertEqual(sheet2.cell(row=3, column=header_map2["Source IP"]).value, "192.168.1.1")
        # Row 3 (risk004, no events)
        self.assertEqual(sheet2.cell(row=4, column=header_map2[str(RiskExportField.RISK_ID.label)]).value, "risk004")
        self.assertEqual(sheet2.cell(row=4, column=header_map2["Source IP"]).value, None)

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    @mock.patch("services.web.risk.handlers.risk_export_service.MailSender")
    @mock.patch("services.web.risk.tasks.export_risks_to_mail.update_state")
    def test_export_risks_to_mail_task_builds_file_and_sends_downloadable_attachment(
        self, mock_update_state, mock_mail_sender, mock_get_event_list, mock_load_authed_risks
    ):
        sender = mock_mail_sender.return_value
        sender.send.return_value = {"result": True}
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)
        mock_get_event_list.return_value = [{"results": [{"event_data": {"user": "alice", "action": "login_success"}}]}]

        result = export_risks_to_mail(
            username="admin",
            risk_ids=[self.risk_1.risk_id],
            risk_view_type=RiskViewType.ALL.value,
            requested_at="2026-06-22 19:33:36",
        )

        self.assertEqual(result["total"], 1)
        mock_update_state.assert_any_call(state="RUNNING", meta={"current": 0, "total": 1})
        mock_update_state.assert_any_call(state="PROGRESS", meta={"current": 1, "total": 1})
        sender.send.assert_called_once()
        attachment = mock_mail_sender.call_args.kwargs["attachments"][0]
        workbook = openpyxl.load_workbook(io.BytesIO(base64.b64decode(attachment["content"])))
        sheet = workbook[self.strategy_1.build_sheet_name()]
        header_map = {cell.value: index for index, cell in enumerate(sheet[1], 1)}
        self.assertEqual(sheet.max_row, 2)
        self.assertEqual(sheet.cell(row=2, column=header_map[str(RiskExportField.RISK_ID.label)]).value, "risk001")
        self.assertEqual(sheet.cell(row=2, column=header_map["Username"]).value, "alice")
        self.assertEqual(sheet.cell(row=2, column=header_map["Action"]).value, "login_success")

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export_scene_view_type_filename(self, mock_get_event_list, mock_load_authed_risks):
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)
        mock_get_event_list.return_value = [{"results": []}]

        resp = resource.risk.risk_export(risk_ids=[self.risk_1.risk_id], risk_view_type=RiskViewType.SCENE.value)
        content_disposition = resp["Content-Disposition"]

        self.assertIn("attachment; filename*", content_disposition)
        self.assertIn("场景风险", unquote(content_disposition))

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    @mock.patch.object(ListEvent, "bulk_request")
    def test_risk_export_queries_events_with_risk_time_range(self, mock_get_event_list, mock_load_authed_risks):
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)
        mock_get_event_list.return_value = [{"results": []}]

        resource.risk.risk_export(risk_ids=[self.risk_1.risk_id])

        params = mock_get_event_list.call_args.args[0][0]
        risk = Risk.objects.get(risk_id=self.risk_1.risk_id)
        expected_start_time = mstimestamp_to_date_string(int(risk.event_time.timestamp() * 1000))
        expected_end_time = mstimestamp_to_date_string(int(risk.event_end_time.timestamp() * 1000))
        self.assertEqual(params["start_time"], expected_start_time)
        self.assertEqual(params["end_time"], expected_end_time)

    @mock.patch("services.web.risk.models.Risk.load_authed_risks")
    def test_risk_export_blocks_unauthed_risks(self, mock_load_authed_risks):
        from services.web.risk.exceptions import ExportRiskNoPermission

        strategy = Strategy.objects.create(
            strategy_id=3,
            strategy_name="Strategy C",
            risk_level=RiskLevel.LOW,
        )
        BindingMetadataHelper.create_resource_binding(
            resource_id=str(strategy.strategy_id),
            resource_type=ResourceVisibilityType.STRATEGY,
            scene_id=self.other_scene.scene_id,
        )
        risk = Risk.objects.create(
            risk_id="risk-outside-scene",
            title="Outside Scene",
            strategy=strategy,
            status=RiskStatus.NEW,
            event_time=datetime.datetime(2023, 1, 5, 10, 0, 0),
            event_end_time=datetime.datetime(2023, 1, 5, 11, 0, 0),
        )
        mock_load_authed_risks.return_value = Risk.objects.filter(risk_id=self.risk_1.risk_id)

        with self.assertRaises(ExportRiskNoPermission):
            resource.risk.risk_export(risk_ids=[risk.risk_id])

    def test_retrieve_strategy_serializer_populates_raw_event_description(self):
        strategy = Strategy(
            strategy_id=3,
            strategy_name="Strategy C",
            risk_level=RiskLevel.LOW,
            event_basic_field_configs=[
                {
                    "field_name": "raw_event_id",
                    "display_name": "raw_event_id",
                    "description": "",
                    "is_priority": True,
                    "duplicate_field": False,
                }
            ],
        )

        serializer = RetrieveRiskStrategyInfoResponseSerializer(instance=strategy)
        data = serializer.data
        raw_config = next(
            cfg for cfg in data.get("event_basic_field_configs", []) if cfg.get("field_name") == "raw_event_id"
        )

        self.assertEqual(raw_config["description"], str(RAW_EVENT_ID_REMARK))

    def test_retrieve_strategy_serializer_returns_scene_id(self):
        strategy = Strategy.objects.create(
            strategy_name="Strategy Scene Serializer",
            risk_level=RiskLevel.LOW,
            event_basic_field_configs=[],
            event_data_field_configs=[],
            event_evidence_field_configs=[],
            risk_meta_field_config=[],
        )
        BindingMetadataHelper.create_resource_binding(
            resource_id=str(strategy.strategy_id),
            resource_type=ResourceVisibilityType.STRATEGY,
            scene_id=self.scene.scene_id,
        )

        serializer = RetrieveRiskStrategyInfoResponseSerializer(instance=strategy)
        data = serializer.data

        self.assertEqual(data["scene_id"], self.scene.scene_id)
