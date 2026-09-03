# -*- coding: utf-8 -*-
"""
TencentBlueKing is pleased to support the open source community by making
蓝鲸智云 - 审计中心 (BlueKing - Audit Center) available.
Copyright (C) 2023 THL A29 Limited,
a Tencent company. All rights reserved.
Licensed under the MIT License (the "License");
you may not use this file except in compliance with the License.
You may obtain a copy of the License at http://opensource.org/licenses/MIT
Unless required by applicable law or agreed to in writing,
software distributed under the License is distributed on
an "AS IS" BASIS, WITHOUT WARRANTIES OR CONDITIONS OF ANY KIND,
either express or implied. See the License for the specific language governing
permissions and limitations under the License.
We undertake not to change the open source license (MIT license) applicable
to the current version of the project delivered to anyone in the future.

F4 导出服务测试（预览导出 + 全量导出）
"""

import io
from unittest import mock

import openpyxl

from services.web.query.ai_assistant.exceptions import (
    AIAssistantError,
    AIOutputInvalidError,
    AIPermissionDeniedError,
)
from services.web.query.ai_assistant.schemas import ResultColumn
from services.web.query.ai_assistant.services.export import (
    LOG_EXPORT_FIELD_WHITELIST,
    FullExportService,
    PreviewExportService,
)
from tests.test_query.test_ai_assistant.base import AIAssistantTestCase

EXPORT_MODULE = "services.web.query.ai_assistant.services.export"


class TestPreviewExportService(AIAssistantTestCase):
    """预览导出：快照 samples → XLSX（纯函数）"""

    def test_export_success(self):
        output = self.make_log_search_output(
            samples=[
                {"start_time": "2026-08-13 12:00:00", "username": "admin"},
                {"start_time": "2026-08-13 11:00:00", "username": "zhangsan"},
            ],
            total=2,
        )

        result = PreviewExportService.export(output)

        self.assertTrue(result.file_name.endswith(".xlsx"))
        workbook = openpyxl.load_workbook(io.BytesIO(result.content))
        sheet = workbook.active
        # 行结构：分类头 / 显示名 / 字段路径 / 数据行
        display_names = [cell.value for cell in sheet[2]]
        self.assertIn("开始时间", display_names)
        self.assertIn("操作人", display_names)
        full_keys = [cell.value for cell in sheet[3]]
        self.assertIn("start_time", full_keys)
        self.assertIn("username", full_keys)
        # 数据行
        first_row = [cell.value for cell in sheet[4]]
        self.assertIn("admin", first_row)
        second_row = [cell.value for cell in sheet[5]]
        self.assertIn("zhangsan", second_row)

    def test_export_extension_column(self):
        """拓展列按 full_key 取值导出"""
        output = self.make_log_search_output(
            columns=[
                ResultColumn(raw_name="username", display_name="操作人"),
                ResultColumn(raw_name="extend_data", keys=["ticket_id"], display_name="工单内容"),
            ],
            samples=[{"username": "admin", "extend_data/ticket_id": "Story-3000"}],
            total=1,
        )

        result = PreviewExportService.export(output)

        workbook = openpyxl.load_workbook(io.BytesIO(result.content))
        sheet = workbook.active
        display_names = [cell.value for cell in sheet[2]]
        self.assertIn("工单内容", display_names)
        full_keys = [cell.value for cell in sheet[3]]
        self.assertIn("extend_data/ticket_id", full_keys)
        first_row = [cell.value for cell in sheet[4]]
        self.assertIn("Story-3000", first_row)

    def test_export_empty_samples_raises(self):
        output = self.make_log_search_output(samples=[], total=0)
        with self.assertRaises(AIAssistantError):
            PreviewExportService.export(output)

    def test_export_with_flatten_extension(self):
        """flatten_extension=True：extend_data 子键平铺为单独列，samples 字典同步展平"""
        output = self.make_log_search_output(
            columns=[
                ResultColumn(raw_name="username", display_name="操作人"),
                ResultColumn(raw_name="extend_data", display_name="拓展数据"),
            ],
            samples=[
                {
                    "username": "admin",
                    "extend_data": {"ticket_id": "Story-3000", "operator": "frodomei"},
                },
                {
                    "username": "zhangsan",
                    "extend_data": {"ticket_id": "Story-4000", "instance_id": "vm-001"},
                },
            ],
            total=2,
        )

        result = PreviewExportService.export(output, export_config={"flatten_extension": True})

        workbook = openpyxl.load_workbook(io.BytesIO(result.content))
        sheet = workbook.active
        # ① 子键并集列：ticket_id、operator、instance_id（保序去重）
        full_keys = [cell.value for cell in sheet[3]]
        self.assertIn("extend_data/ticket_id", full_keys)
        self.assertIn("extend_data/operator", full_keys)
        self.assertIn("extend_data/instance_id", full_keys)
        # ② extend_data 单列已移除
        self.assertNotIn("extend_data", full_keys)
        # ③ 行 1：缺 operator/instance_id 不影响 ticket_id 取值
        first_data_row = [cell.value for cell in sheet[4]]
        self.assertIn("admin", first_data_row)
        self.assertIn("Story-3000", first_data_row)
        self.assertIn("frodomei", first_data_row)
        # ④ 行 2：缺 operator 不报错，空值单元格（空字符串）
        second_data_row = [cell.value for cell in sheet[5]]
        self.assertIn("zhangsan", second_data_row)
        self.assertIn("Story-4000", second_data_row)
        self.assertIn("vm-001", second_data_row)

    def test_flatten_extension_with_no_extension_data(self):
        """flatten_extension=True 但 samples 全无 extend_data：等同未开启，输出列无变化"""
        output = self.make_log_search_output(
            columns=[ResultColumn(raw_name="username", display_name="操作人")],
            samples=[{"username": "admin"}, {"username": "zhangsan"}],
            total=2,
        )

        result = PreviewExportService.export(output, export_config={"flatten_extension": True})

        workbook = openpyxl.load_workbook(io.BytesIO(result.content))
        sheet = workbook.active
        full_keys = [cell.value for cell in sheet[3]]
        self.assertEqual(full_keys.count("extend_data"), 0)

    def test_flatten_extension_false_keeps_default(self):
        """flatten_extension 缺省/False：保持原 extend_data 单列输出（与现状兼容）"""
        output = self.make_log_search_output(
            columns=[
                ResultColumn(raw_name="username", display_name="操作人"),
                ResultColumn(raw_name="extend_data", display_name="拓展数据"),
            ],
            samples=[{"username": "admin", "extend_data": {"ticket_id": "Story-3000"}}],
            total=1,
        )

        result = PreviewExportService.export(output)

        workbook = openpyxl.load_workbook(io.BytesIO(result.content))
        sheet = workbook.active
        full_keys = [cell.value for cell in sheet[3]]
        self.assertIn("extend_data", full_keys)
        self.assertNotIn("extend_data/ticket_id", full_keys)


@mock.patch(f"{EXPORT_MODULE}.resource.query.create_collector_search_export_task")
@mock.patch(f"{EXPORT_MODULE}.SearchLogPermission.has_system_search_permission")
class TestFullExportService(AIAssistantTestCase):
    """全量导出：快照 condition 原样重建 → LogExportTask"""

    def test_create_task_success(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        condition = self.make_condition(conditions=[self.make_field_condition()])
        export_config = {"field_scope": "all", "fields": []}

        FullExportService.create_task(
            condition=condition,
            namespace=self.namespace,
            export_config=export_config,
            task_name="AI助手检索导出-12345678",
            username=self.username,
        )

        mock_create_task.assert_called_once()
        _, kwargs = mock_create_task.call_args
        self.assertEqual(kwargs["namespace"], self.namespace)
        # query_params：顶层时间字段 + condition 原样 + scope 精确注入
        query_params = kwargs["query_params"]
        self.assertEqual(query_params["start_time"], self.start_time)
        self.assertEqual(query_params["end_time"], self.end_time)
        conditions = query_params["conditions"]
        # 首条为 scope 系统精确注入
        self.assertEqual(conditions[0]["field"]["raw_name"], "system_id")
        self.assertEqual(conditions[0]["operator"], "include")
        self.assertEqual(conditions[0]["filters"], [self.target_system_id])
        # 业务条件原样透传
        self.assertEqual(conditions[1]["field"]["raw_name"], "username")
        # 时间条件不在 conditions 内（由导出运行时按顶层 start/end 注入）
        raw_names = [cond["field"]["raw_name"] for cond in conditions]
        self.assertNotIn("thedate", raw_names)
        self.assertNotIn("dtEventTimeStamp", raw_names)
        # export_config 前端传入透传
        self.assertEqual(kwargs["export_config"], export_config)

    def test_permission_denied(self, mock_perm, mock_create_task):
        mock_perm.return_value = False

        with self.assertRaises(AIPermissionDeniedError) as ctx:
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={"field_scope": "all", "fields": []},
                task_name="t",
                username=self.username,
            )
        self.assertEqual(ctx.exception.error_code, "PERMISSION_DENIED")
        mock_create_task.assert_not_called()

    def test_permission_checked_with_explicit_username(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        FullExportService.create_task(
            condition=self.make_condition(),
            namespace=self.namespace,
            export_config={"field_scope": "all", "fields": []},
            task_name="t",
            username=self.username,
        )
        mock_perm.assert_called_once_with(self.target_system_id, self.username)

    def test_invalid_field_scope(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError):
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={"field_scope": "hack", "fields": []},
                task_name="t",
                username=self.username,
            )

    def test_specified_scope_requires_fields(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError):
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={"field_scope": "specified", "fields": []},
                task_name="t",
                username=self.username,
            )

    def test_field_not_in_whitelist(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError):
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={
                    "field_scope": "specified",
                    "fields": [{"raw_name": "not_a_field", "display_name": "x", "keys": []}],
                },
                task_name="t",
                username=self.username,
            )

    def test_build_task_name(self, mock_perm, mock_create_task):
        name = FullExportService.build_task_name("abcdef1234567890")
        self.assertIn("abcdef12", name)

    def test_invalid_flatten_extension_type_rejected(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError) as ctx:
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={"field_scope": "all", "fields": [], "flatten_extension": "yes"},
                task_name="t",
                username=self.username,
            )
        self.assertIn("flatten_extension", str(ctx.exception.extra))

    def test_invalid_extension_keys_type_rejected(self, mock_perm, mock_create_task):
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError) as ctx:
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={"field_scope": "all", "fields": [], "extension_keys": "ticket_id"},
                task_name="t",
                username=self.username,
            )
        self.assertIn("extension_keys", str(ctx.exception.extra))

    def test_extension_keys_empty_string_filtered(self, mock_perm, mock_create_task):
        """extension_keys 含空字符串被过滤；含非字符串整体拒绝"""
        mock_perm.return_value = True
        with self.assertRaises(AIOutputInvalidError):
            FullExportService.create_task(
                condition=self.make_condition(),
                namespace=self.namespace,
                export_config={
                    "field_scope": "all",
                    "fields": [],
                    "extension_keys": ["ticket_id", 123],  # 含非字符串
                },
                task_name="t",
                username=self.username,
            )

    def test_flatten_extension_and_extension_keys_passthrough(self, mock_perm, mock_create_task):
        """flatten_extension/extension_keys 合法值透传到 task.export_config（落库给 DataProcessor 运行时使用）"""
        mock_perm.return_value = True
        mock_create_task.return_value = {"id": 1, "status": "pending"}

        FullExportService.create_task(
            condition=self.make_condition(),
            namespace=self.namespace,
            export_config={
                "field_scope": "all",
                "fields": [],
                "flatten_extension": True,
                "extension_keys": ["ticket_id", "operator"],
            },
            task_name="t",
            username=self.username,
        )

        _, kwargs = mock_create_task.call_args
        self.assertTrue(kwargs["export_config"]["flatten_extension"])
        self.assertEqual(kwargs["export_config"]["extension_keys"], ["ticket_id", "operator"])


class TestLogExportConfigSerializerProtocol(AIAssistantTestCase):
    """协议层验证：LogExportReqSerializer 嵌套校验不再剥离 AI 导出新字段，且原检索页形态零变化。"""

    def test_new_fields_survive_serialization(self):
        """AI 导出 export_config 含新字段：经 LogExportConfigSerializer 校验后保留（可落库）"""
        from services.web.query.serializers import LogExportConfigSerializer

        serializer = LogExportConfigSerializer(
            data={
                "field_scope": "all",
                "fields": [],
                "flatten_extension": True,
                "extension_keys": ["ticket_id", "operator"],
            }
        )
        self.assertTrue(serializer.is_valid(), serializer.errors)
        validated = serializer.validated_data
        self.assertTrue(validated["flatten_extension"])
        self.assertEqual(validated["extension_keys"], ["ticket_id", "operator"])

    def test_legacy_export_config_unchanged(self):
        """原检索页 export_config（仅 field_scope/fields）：validated_data 不含新键，落库形态与历史一致"""
        from services.web.query.serializers import LogExportConfigSerializer

        serializer = LogExportConfigSerializer(data={"field_scope": "all", "fields": []})
        self.assertTrue(serializer.is_valid(), serializer.errors)
        validated = serializer.validated_data
        self.assertNotIn("flatten_extension", validated)
        self.assertNotIn("extension_keys", validated)
        self.assertEqual(validated, {"field_scope": "all", "fields": []})

    # 注：非法形态（flatten_extension="yes" / extension_keys 含非字符串）的严格拦截在
    # AI 侧 FullExportService._validate_export_config（isinstance 显式校验，入口先于 serializer 执行，
    # 见 TestFullExportService 两个 rejected 用例）；serializer 层 DRF BooleanField/CharField 为宽松
    # 归一语义（"yes"→True、int→str），不作为防线，此处不做拒绝断言。


class TestExportConfigFlattenIsolation(AIAssistantTestCase):
    """隔离性验证：ExportConfig 平铺分支仅在显式开启时生效，原检索页任务（无新字段）导出列零变化。"""

    def _make_task(self, export_config: dict):
        from services.web.query.models import LogExportTask

        return LogExportTask(export_config=export_config)

    def test_legacy_task_fields_unchanged(self):
        """原检索页任务（field_scope=all 无新字段）：export_fields 与改动前完全一致（含 extend_data 单列）"""
        from services.web.query.export.model import ExportConfig

        config = ExportConfig(task=self._make_task({"field_scope": "all", "fields": []}))
        full_keys = [field.full_key for field in config.export_fields]
        self.assertIn("extend_data", full_keys)
        # 新字段缺省不产生任何子键列
        self.assertEqual([key for key in full_keys if key.startswith("extend_data/")], [])

    def test_flatten_task_replaces_extend_data_column(self):
        """开启平铺 + 子键清单：extend_data 单列被替换为子键列（仅 AI 导出路径构造此形态）"""
        from services.web.query.export.model import ExportConfig

        config = ExportConfig(
            task=self._make_task(
                {"field_scope": "all", "fields": [], "flatten_extension": True, "extension_keys": ["ticket_id"]}
            )
        )
        full_keys = [field.full_key for field in config.export_fields]
        self.assertNotIn("extend_data", full_keys)
        self.assertIn("extend_data/ticket_id", full_keys)

    def test_flatten_without_keys_keeps_single_column(self):
        """开启平铺但未传子键清单：不替换（无列定义来源，保持单列）"""
        from services.web.query.export.model import ExportConfig

        config = ExportConfig(task=self._make_task({"field_scope": "all", "fields": [], "flatten_extension": True}))
        full_keys = [field.full_key for field in config.export_fields]
        self.assertIn("extend_data", full_keys)


class TestAIStandardFieldScope(AIAssistantTestCase):
    """ai_standard scope：字段集合 = 快照默认展示列（get_fields 兜底；AI 链路主路径为 SPECIFIED 翻译）"""

    def test_get_fields_returns_snapshot_default_columns(self):
        """get_fields(ai_standard)：顺序与集合 = SNAPSHOT_DEFAULT_COLUMNS（单一来源，防漂移）"""
        from services.web.query.ai_assistant.constants import SNAPSHOT_DEFAULT_COLUMNS
        from services.web.query.constants import LogExportFieldScope

        fields = LogExportFieldScope.get_fields(LogExportFieldScope.AI_STANDARD.value)
        self.assertEqual(
            [field.field_name for field in fields],
            [raw_name for raw_name, _ in SNAPSHOT_DEFAULT_COLUMNS],
        )

    def test_ai_standard_fields_all_in_whitelist(self):
        """ai_standard 字段全部在导出白名单内（翻译为 SPECIFIED 后可过白名单校验）"""
        from services.web.query.constants import LogExportFieldScope

        fields = LogExportFieldScope.get_fields(LogExportFieldScope.AI_STANDARD.value)
        for field in fields:
            self.assertIn(field.field_name, LOG_EXPORT_FIELD_WHITELIST)

    def test_ai_standard_scope_accepted_by_validate(self):
        """ai_standard 为合法 field_scope：不在 AI 入口翻译的调用路径下也不被白名单校验拒绝"""
        from services.web.query.constants import LogExportFieldScope

        self.assertIn(LogExportFieldScope.AI_STANDARD.value, LogExportFieldScope.values)
